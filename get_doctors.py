from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import selenium
#from selenium.webdriver.support.ui import Select
import openpyxl as xl
from bs4 import BeautifulSoup as bss
import os
import re
from urllib.parse import urlparse
import time
#***********configures****************
browser_platform = 'firefox'

#*************************************
#*********pre handling
headers={
    'basic_info':['Doc_Name','Dept','Title','Expertise','Bio','Thank_Letter',
            'Gift','Badge_2003','Badge_2004','Badge_2005','Badge_2006','Badge_2007',
            'Badge_2008','Badge_2009','Badge_2010','Badge_2011','Badge_2012','Badge_2013',
            'Badge_2014','Badge_2015','Badge_2016','Personal_Site','Ratings','Satisfaction_Rate',
            'Accu_patient','Attitude_Rate','Recent_2wks','exp_kidneystone','treated_patients',
            'following_patient','afterDiag_ratings','votes_kidneyStone','2yrs_Vote_Effective',
            '2yrs_Vote_Attitude','red','green','blue'],
    'experience':['doc','disease','patient','geo_loc','source','time','content','doc_reply'],
    'gift':['doc','type','time','id'],
    'post':['doc','title','read','date','contribution_score','total_visit','#post','#patient',
           'outpatientvisit_yesterday','wechat','total_after','total_votes',
           'lastseen','private_webopen'],
    'primarycare': ['doc','available','mon','3mon','6mon','12mon'],
    'service':['doc','patient','title','justpaid','after','tele','thank','gift',
            'special_gift','encrypted','disease','d','p','last_post'],
    'sharing':['doc','type','patient','disease','aim','treatment_type','outcome','attitude','time','method',
            'content','reason','channel','status','expense','useful','#recommendation','id'],
    'tele_review':['doc','patient','#use','time','content','tag1','tag2','tag3','tag4','tag5','tag6','tag7'],
    'tele':['doc','price','option1','option2','this_month','review','good','fair','bad','tag1','tag2','tag3',
    'tag4','tag5','tag6','tag7']
}
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
root_url = 'http://www.haodf.com/jibing/shenjieshi/daifu_{}_beijing_all_all_all.htm'
log_dir = 'log'
results_dir = 'results'
loc_dir = 'local'
list_personal_url = log_dir + '/list_personal_url.txt'
list_exp_url = log_dir + '/list_exp_url.txt'
list_share_url = log_dir + '/list_share_url.txt'
list_gift_url = log_dir + '/list_gift_url.txt'
list_got_gift_url = log_dir + '/list_got_gift_url.txt'
list_got_exp_url = log_dir + '/list_got_exp_url.txt'
list_got_share_url = log_dir + '/list_got_share_url.txt'
pwd = os.getcwd()
if not os.path.isdir(log_dir):
    os.makedirs(log_dir)
if not os.path.isdir(results_dir):
    os.makedirs(results_dir)

basic_info = results_dir + '/basic_info.xlsx'
experience = results_dir + '/experience.xlsx'
gift = results_dir + '/gift.xlsx'
post = results_dir + '/post.xlsx'
primarycare = results_dir + '/primarycare.xlsx'
service = results_dir + '/service.xlsx'
sharing = results_dir + '/sharing.xlsx'
tele = results_dir + '/tele.xlsx'
tele_review = results_dir + '/tele_review.xlsx'
def create_if_not_exist(excel_file):
    if not os.path.isfile(excel_file):
        pure_file_name = re.sub('^.*./','',excel_file)
        pure_file_name = re.sub('\..*$','',pure_file_name)
        excel_header = headers[pure_file_name]
        wb = xl.Workbook()
        sheet = wb.active
        sheet.append(excel_header)
        wb.save(excel_file)
        print('created {}'.format(excel_file))

for excel_file in [basic_info,experience,gift,post,primarycare,
    service,sharing,tele,tele_review]:
    create_if_not_exist(excel_file)


browser = ''
def start_browser(allow_img=False):
    global browser
    if browser_platform == "chrome":
        if allow_img:
            browser = webdriver.Chrome()
        else:
            chromeOptions = webdriver.ChromeOptions()
            prefs = {"profile.managed_default_content_settings.images":2}
            chromeOptions.add_experimental_option("prefs",prefs)
            browser = webdriver.Chrome(chrome_options=chromeOptions)
    elif browser_platform == "firefox":
        firefox_profile = webdriver.FirefoxProfile()
        # Disable CSS
        firefox_profile.set_preference('permissions.default.stylesheet', 2)
        # Disable images
        if allow_img == False:
            firefox_profile.set_preference('permissions.default.image', 2)

        # Disable Flash
        #firefox_profile.set_preference('dom.ipc.plugins.enabled.libflashplayer.so', 'false')
        # Set the modified profile while creating the browser object
        browser = webdriver.Firefox(firefox_profile=firefox_profile)
        #browser = webdriver.Firefox()
    else:
        if allow_img == False:
            browser = webdriver.PhantomJS(service_args=['--load-images=false'])
        else:
            browser = webdriver.PhantomJS()
    return browser
browser = start_browser()
basic_info_got = log_dir + '/basic_info.txt'
#**********



def element_exists(element,xpath):
    try:
        element.find_element_by_xpath(xpath)
        return 1
    except:
        return 0
def set_browser(url,timeout=60):
    browser.set_page_load_timeout(timeout)
    try:
        browser.get(url)
    except:
        pass
def set_browser_local(url,replace=False,js=False,timeout=None):
    filename = parse_url_to_local_file(url)
    local_path = "file://" +pwd +'/'+ filename
    if os.path.isfile(filename) and replace == False:
        browser.get(local_path)
    else:
        download_site(url,replace,js,timeout)
        browser.get(local_path)

def write_log(msg,file='log/log.txt'):
    print(msg)
    with open(file,'a') as f:
        f.write(msg+'\n')

def parse_url_to_local_file(url):
    o = urlparse(url)
    filename =  loc_dir + '/' + o.netloc + o.path
    if o.query:
        suffixm = re.search('\.[^.]*$',filename)
        if suffixm:
            suffix = suffixm.group()
        else:
            suffix = ''
        file_nosuff = re.sub('\.[^.]*$','',filename)
        filename = file_nosuff + '_' + o.query + suffix
    filename = re.sub('/$','',filename)
    return filename

def download_site(url,replace=True,js=False,timeout=None):
    filename = parse_url_to_local_file(url)
    if replace == False and os.path.isfile(filename):
        return
    dirname = os.path.dirname(filename)
    if not os.path.isdir(dirname):
        os.makedirs(dirname)
    if timeout:
        browser.set_page_load_timeout(timeout)
    try:
        browser.get(url)
    except:
        print('timeout, try continue with following steps')
        pass
    source = browser.page_source
    source = re.sub('<meta charset=[^<>]*>','<meta charset="utf8" />',source)
    source = re.sub('charset=gb2312','charset=utf8',source)
    bs_page = bss(source,'html.parser')
    if not js:
        scripts = bs_page.find_all('script')
        for script in scripts:
            script.extract()
    links = bs_page.find_all('link')
    for link in links:
        link.extract()
    with open(filename,'w') as f:
        f.write(bs_page.prettify())
    write_log('{} downloaded to local'.format(url))

def get_basic_info(page_num,doc_num,replace=False):
    '''
    get all the number of page "page_num" and doctor "doc_num".
    For example,
    get_all_info(1,1) will get the info of doctor 叶雄俊
    get_all_info(1,2) will get the info of doctor 胡卫国
    get_all_info(2,1) will get the info of doctor 杨波
    '''
    page_url = root_url.format(page_num)
    try:
        set_browser_local(page_url,replace)
    except:
        pass
    doctor_infos = browser.find_elements_by_xpath('//ul[@class="fs hp_doc clearfix"]/li[@class="hp_doc_box_serviceStar"]')
    if len(doctor_infos) < doc_num:
        print("for page {}, doc_num should < {}".format(page_num,len(doctor_infos)))
        return 1
    doctor_info = doctor_infos[doc_num-1]
    a = doctor_info.find_element_by_xpath('.//div[@class="oh zoom lh180"]//a[@class="blue_a3"]')
    try:
        a_personal = doctor_info.find_element_by_xpath('.//a[contains(text(),"个人网站")]')
        Personal_Site = a_personal.get_attribute('href')
    except:
        Personal_Site = ""
    Doc_Name = a.text.strip()
    red = element_exists(doctor_info,".//a[@class='tel_btn']")
    green = element_exists(doctor_info,".//a[@class='online_btn']")
    blue = element_exists(doctor_info,".//a[@class='jiahao_btn']")
    try:
        set_browser_local(a.get_attribute('href'),replace)
    except:
        pass
    intro_element = browser.find_element_by_xpath("//div[@class='lt']/table/tbody")
    badges = [0]*14
    for year in range(2003,2017):
        #xpath = ".//img[@src='http://i1.hdfimg.com/doctorzone/images/doctorhonour/{}/selfgooddoctor_{}.png']".format(year,year)
        xpath1 = ".//a[@title='{}年度好大夫']".format(year)
        xpath2 = ".//a[@title='{}年度好大夫 ']".format(year)
        if element_exists(intro_element,xpath1) ==1 or element_exists(intro_element,xpath2)==1:
            badges[year-2003] = 1
    dept_tr = intro_element.find_element_by_xpath('.//td[contains(text(),"科　　室：")]')
    dept_tr = dept_tr.find_element_by_xpath('..')
    Dept = dept_tr.find_element_by_xpath('./td[3]').text.strip()
    try:
        Gift = dept_tr.find_element_by_xpath('.//a[contains(text(),"礼物")]/span').text.strip()
        Gift = re.sub('[^0-9]*','',Gift)
    except:
        Gift = ''
    try:
        Thank_Letter = dept_tr.find_element_by_xpath('.//a[contains(text(),"感谢信")]/span').text.strip()
        Thank_Letter = re.sub('[^0-9]*','',Thank_Letter)
    except:
        Thank_Letter = ''
    title_tr = intro_element.find_element_by_xpath('.//td[contains(text(),"职　　称：")]')
    title_tr = title_tr.find_element_by_xpath('..')
    Title = title_tr.find_element_by_xpath('./td[3]').text.strip()
    try:
        Expertise = browser.find_element_by_xpath('//div[@id="full_DoctorSpecialize"]').get_attribute('innerHTML')
    except:
        Expertise = ''
    Expertise = Expertise.strip()
    Expertise = re.sub('<span>.*$','',Expertise).strip()
    Expertise = re.sub('<[^<>]*>','',Expertise).strip()
    Expertise = re.sub('<<收起','',Expertise)
    if re.search('^暂无',Expertise):
        Expertise = ''
    try:
        Bio = browser.find_element_by_xpath('//div[@id="full"]').get_attribute('innerHTML')
    except:
        Bio = browser.find_element_by_xpath('.//td[contains(text(),"执业经历：")]')
        Bio = Bio.find_element_by_xpath('..')
        Bio = Bio.find_element_by_xpath('./td[3]').text.strip()
    Bio = Bio.strip()
    Bio = re.sub('<span>.*$','',Bio).strip()
    Bio = re.sub('<br>','',Bio)
    Bio = re.sub('[\n]+','\n',Bio)
    Bio = re.sub('<[^<>]*>','',Bio)
    Bio = re.sub('<<收起','',Bio)
    if re.search('^暂无',Bio):
        Bio = ''
    Ratings = browser.find_element_by_xpath('//div[@class="recommend-part"]//p[@class="r-p-l-score"]').text.strip()
    spans = browser.find_elements_by_xpath('//div[@class="recommend-part"]//div[@class="fl score-part"]//span')
    dic_span = {}
    for span in spans:
        text = span.text.split('：')
        dic_span[text[0].strip()] = text[1].strip()
    Satisfaction_Rate,Accu_patient,Attitude_Rate,Recent_2wks = '','','',''
    if '疗效满意度' in dic_span:
        Satisfaction_Rate = dic_span['疗效满意度']
    if '累计帮助患者数' in dic_span:
        Accu_patient = dic_span['累计帮助患者数']
    if '态度满意度' in dic_span:
        Attitude_Rate = dic_span['态度满意度']
    if '近两周帮助患者数' in dic_span:
        Recent_2wks = dic_span['近两周帮助患者数']
    try:
        td = browser.find_element_by_xpath('//td[contains(text(),"诊治过的患者数")]')
        treated_patients = td.text.split('：')[1].strip().replace('例','')
    except:
        treated_patients = ''
    try:
        td = browser.find_element_by_xpath('//td[contains(text(),"随访中的患者数")]')
        following_patient = td.text.split('：')[1].strip().replace('例','')
    except:
        following_patient = ''
    img_lights=browser.find_elements_by_xpath("//td[@class='fuwuStarImg']//img[@src='http://i1.hdfimg.com/www/images/starRightliang.png']")
    afterDiag_ratings = len(img_lights)
    try:
        effect_text = browser.find_element_by_xpath('//td[contains(text(),"疗效")]').find_element_by_xpath('..').text
        effect_vote = re.findall('[0-9%]+',effect_text)[0]
    except:
        effect_vote = ''
    try:
        attitude_text = browser.find_element_by_xpath('//td[contains(text(),"态度")]').find_element_by_xpath('..').text
        attitude_vote = re.findall('[0-9%]+',attitude_text)[0]
    except:
        attitude_vote = ''
    try:
        div_experience = browser.find_element_by_xpath('//div[@id="bp_doctor_servicestar"]')
    except:
        div_experience = ""
    exp_url = ''
    exp_kidneystone = ''
    if div_experience:
        texts = re.findall('肾结石[^0-9]*([0-9]+)例',div_experience.text)
        if texts:
            exp_kidneystone = texts[0]
        else:
            try:
                more_a = div_experience.find_element_by_xpath('.//a[contains(text(),"查看详情>>")]')
                exp_url = more_a.get_attribute('href')
            except:
                pass
    try:
        div_good = browser.find_element_by_xpath('//div[@id="bp_doctor_getvote"]')
    except:
        div_good = ''
    votes_kidneyStone = ''
    if div_good:
        texts = re.findall('肾结石[^0-9]*([0-9]+)票',div_good.text)
        if texts:
            votes_kidneyStone = texts[0]
    return ([Doc_Name,Dept,Title,Expertise,Bio,Thank_Letter,Gift] +
            badges + [Personal_Site,Ratings,Satisfaction_Rate,Accu_patient,Attitude_Rate] +
            [Recent_2wks,exp_kidneystone,treated_patients,following_patient,
            afterDiag_ratings,votes_kidneyStone,
            effect_vote,attitude_vote,red,green,blue])
def get_exp_kidney(url):
    if not url:
        return ''
    set_browser(url)
    div = browser.find_element_by_xpath('//div[@id="tabmainin"]')
    texts = re.findall('肾结石[^0-9]*([0-9]+)例',div.text)
    if texts:
        return texts[0]
    else:
        return ''

def load_got_basic_info():
    dic = set()
    if not os.path.isfile(basic_info_got):
        return dic
    with open(basic_info_got) as f:
        for line in f.readlines():
            line = line.strip()
            lines = line.split(',')
            x,y = [int(z) for z in lines]
            dic.add((x,y))
    return dic
def write_got_basic_info(page_num,doc_num):
    with open(basic_info_got,'a') as f:
        f.write("{},{}\n".format(page_num,doc_num))
def get_store_basic_info(page_num,doc_num):
    dic = load_got_basic_info()
    if (page_num,doc_num) in dic:
        print("basic info of {},{} already got".format(page_num,doc_num))
        return
    line = get_basic_info(page_num,doc_num)
    if line==1:
        line = get_basic_info(page_num,doc_num,replace=True)
    wb = xl.load_workbook(basic_info)
    sheet = wb.active
    sheet.append(line)
    wb.save(basic_info)
    write_got_basic_info(page_num,doc_num)
    write_log('get basic info of {},{}'.format(page_num,doc_num))
#for page_num in range(1,23):
#    for doc_num in range(1,16):
#        get_store_basic_info(page_num,doc_num)
##get_store_basic_info(4,5)
##download_site(root_url.format(1),True)
##set_browser_local(root_url.format(1))
##line=get_basic_info(20,3)
#print(line)

###Download sites
def download_source(page_num):
    download_site(root_url.format(page_num))


def get_experience_url(url):
    if not url:
        return ''
    set_browser_local(url)
    try:
        div = browser.find_element_by_xpath('//div[@id="tabmainin"]')
        a = div.find_element_by_xpath('.//a[contains(text(),"肾结石")]').get_attribute('href')
    except:
        a = ''
    return a

def get_element(element,xpath):
    try:
        x = element.find_element_by_xpath(xpath)
    except:
        x = ''
    return x
def write_to_list(list_file,url):
    if url:
        with open(list_file,'a') as f:
            f.write(url+'\n')
def load_list(list_file):
    urls = set()
    if not os.path.isfile(list_file):
        return urls
    with open(list_file) as f:
        for line in f.readlines():
            urls.add(line.strip())
    return urls
def download_doc_sites(page_num,doc_num,replace=False):
    #personal sites
    #
    page_url = root_url.format(page_num)
    set_browser_local(page_url)
    doctor_infos = browser.find_elements_by_xpath('//ul[@class="fs hp_doc clearfix"]/li[@class="hp_doc_box_serviceStar"]')
    doctor_info = doctor_infos[doc_num-1]
    try:
        a_personal = doctor_info.find_element_by_xpath('.//a[contains(text(),"个人网站")]')
        Personal_Site = a_personal.get_attribute('href')
    except:
        Personal_Site = ""

    a = doctor_info.find_element_by_xpath('.//div[@class="oh zoom lh180"]//a[@class="blue_a3"]')
    basic_url = a.get_attribute('href')
    set_browser_local(basic_url)
    div_experience = get_element(browser,'//div[@id="bp_doctor_servicestar"]')
    if div_experience:
        try:
            experience_url = div_experience.find_element_by_xpath('.//a[contains(text(),"肾结石")]').get_attribute('href')
        except:
            try:
                more_url = div_experience.find_element_by_xpath('.//a[contains(text(),"查看详情>>")]').get_attribute('href')
                current_url = browser.current_url
                experience_url = get_experience_url(more_url)
                set_browser(current_url)
            except:
                experience_url = ''
    try:
        gift_url = browser.find_element_by_xpath('//div[@class="nav2"]/a[contains(text(),"礼物")]').get_attribute('href')
        if gift_url.find('api/present/ajaxsendtomydoctor?uname=')>=0:
            gift_url = ''
    except:
        gift_url = ''

    share_url_a = get_element(browser,'//div[@class="nav2"]/a[contains(text(),"看病经验")]')
    if share_url_a:
        share_url = share_url_a.get_attribute('href').replace('kanbingjingyan','all')
    else:
        share_url_a = get_element(browser,'//div[@class="nav2"]/a[contains(text(),"感谢信")]')
        if share_url_a:
            share_url = share_url_a.get_attribute('href').replace('ganxiexin','all')
        else:
            share_url_a = get_element(browser,"//table[@class='lbjg']//a[contains(text(),'查看全部')]")
            if share_url:
                share_url = share_url_a.get_attribute('href')
            else:
                share_url = basic_url
    write_to_list(list_personal_url,Personal_Site)
    write_to_list(list_exp_url,experience_url)
    write_to_list(list_gift_url,gift_url)
    write_to_list(list_share_url,share_url)
    if Personal_Site:
        download_site(Personal_Site,replace)
    return (Personal_Site,experience_url,gift_url,share_url)
for page_num in range(1,23):
    for doc_num in range(1,16):
        download_doc_sites(page_num,doc_num)
def get_gifts_one_page():
    div_gift = browser.find_element_by_xpath("//div[@id='comment_content'][@class='doctorjyjy']")
    gift_tables = div_gift.find_elements_by_xpath('.//table[@class="doctorjy"]')
    gifts = []
    doc = get_element(browser,"//div[@id='doctor_header']//h1").text.strip()
    N = len(gift_tables)
    print('N={}'.format(N))
    i = 0
    while i < N:
        try:
            gift_tables = div_gift.find_elements_by_xpath('.//table[@class="doctorjy"]')
            N = len(gift_tables)
            print('N={}'.format(N))
            gift_table = gift_tables[i]
            td = get_element(gift_table,".//td[contains(text(),'患者：')]")
            patient = re.sub('患者：','',td.text).strip() if td else ''
            td = get_element(gift_table,".//td[contains(text(),'时间：')]")
            time = re.sub('时间：','',td.text).strip() if td else ''
            img = get_element(gift_table,".//div[@class='pr']//img")
            img = img.get_attribute('src') if img else ''
            i += 1
        except selenium.common.exceptions.StaleElementReferenceException:
            continue
        gifts.append([doc,img,time,patient])
    return gifts
def get_gifts_one_page_bs():
    bs_page = bss(browser.page_source,'html.parser')
    bs_tables = bs_page.find_all('table',class_='doctorjy')
    if len(bs_tables) < 20:
        time.sleep(1)
        bs_page = bss(browser.page_source,'html.parser')
        bs_tables = bs_page.find_all('table',class_='doctorjy')
    gifts = []
    for bs_table in bs_tables:
        bs_tds = bs_table.find('table').find_all('td')
        bs_td_patient = ''
        bs_td_time = ''
        bs_td_doc = ''
        for bs_td in bs_tds:
            if not bs_td_patient and re.search('患者：',bs_td.text):
                bs_td_patient = bs_td
            if not bs_td_time and re.search('时间：',bs_td.text):
                bs_td_time = bs_td
            if not bs_td_doc and re.search('就诊大夫：',bs_td.text):
                bs_td_doc = bs_td
        patient = re.sub('患者：','',bs_td_patient.text).strip() if bs_td_patient else ''
        ttime = re.sub('时间：','',bs_td_time.text).strip() if bs_td_time else ''
        doc = re.sub('就诊大夫：','',bs_td_doc.text).strip() if bs_td_doc else ''
        bs_img = bs_table.find('div',class_='pr').find('img')
        img = bs_img['src'] if bs_img else ''
        gifts.append([doc,img,ttime,patient])
    return gifts

def get_gifts_one_url(gift_url):
    set_browser(gift_url)
    browser.find_element_by_xpath('//div[@id="present_button"]//a[@rel="nofollow"]').click()
    time.sleep(0.5)
    doc_web = get_element(browser,"//div[@id='doctor_header']//h1").find_element_by_xpath('..').get_attribute('href')
    doc_web = re.sub('\.htm.*$','',doc_web)+'/'
    try:
        total_page_a = get_element(browser,'//div[@class="p_bar"]/a[@class="p_text"][@rel="true"][contains(text(),"共")]')
    #total_page_a = browser.find_element_by_xpath('//div[@class="p_bar"]/a[@class="p_text"][contains(text(),"共")]')
        total_page = total_page_a.text.replace('共','').replace('页','').strip()
        total_page = int(total_page)
    except:
        total_page = ''
    gifts = []
    gift_dic = set()
    for gift in get_gifts_one_page_bs():
        if tuple(gift) not in gift_dic:
            gifts.append(gift)
            gift_dic.add(tuple(gift))

    if total_page and total_page > 1:
        page = 2
        load_count = 0
        while page <= total_page:
            try:
                input_area = get_element(browser,'//div[@class="p_bar"]/input')
                input_area.send_keys(page)
                go_button = get_element(browser,'//div[@class="p_bar"]/button')
                go_button.click()
                time.sleep(1)
            except AttributeError:
                load_count += 1
                if load_count >= 5:
                    browser.find_element_by_xpath('//div[@id="present_button"]//a[@rel="nofollow"]').click()
                    time.sleep(0.5)
                continue
            except selenium.common.exceptions.StaleElementReferenceException:
                continue
            page_gifts =  get_gifts_one_page_bs()
            gifts += page_gifts
            print('page {}'.format(page))
            page+=1
    return gifts
#browser.close()
#browser = start_browser(True)

def get_gifts():
    got_gift_urls = load_list(list_got_gift_url)
    with open(list_gift_url) as f:
        for line in f.readlines():
            line = line.strip()
            if line in got_gift_urls:
                print('{} already parsed'.format(line))
            else:
                gifts = get_gifts_one_url(line)
                wb = xl.load_workbook(gift)
                sheet = wb.active
                with open('results/gifts.txt','a') as f:
                    for gi in gifts:
                        sheet.append(gi)
                        f.write('|'.join(gi)+'\n')
                wb.save(gift)
                write_to_list(list_got_gift_url,line)
                write_log("get {} gifts for {}".format(len(gifts),line))
#get_gifts()
def get_exps_one_page_bs():
    bs_page = bss(browser.page_source,'html.parser')
    bs_divs = bs_page.find_all('div',class_='singlePaitentR clearfix')
    doc = bs_page.find('div',id='doctor_header').find('h1').text.strip()
    exps = []
    for bs_div in bs_divs:
        patient = ''
        geo_loc = ''
        disease = ''
        source = ''
        exp_time = ''
        content = ''
        doc_reply = '0'
        bs_p = bs_div.find('p',class_='starBottom10')
        text = bs_p.text.strip()
        texts = [s.strip() for s in text.split('\n')]
        for text in texts:
            if re.search('患者：',text):
                patient = re.sub('患者：','',text).strip()
            elif re.search('来自：',text):
                geo_loc = re.sub('来自：','',text).strip()
            elif re.search('时间：',text):
                exp_time = re.sub('时间：','',text).strip()
            else:
                source = text.strip()
        bs_span = bs_div.find('span',class_='disName')
        if bs_span:
            disease = re.sub('疾病：','',bs_span.text).strip()
        bs_a = bs_div.find('a',id='refFromServiceStar_cnzz')
        if bs_a:
            doc_reply = '1'
        bs_content = bs_div.find('p',class_='contN')
        content = bs_content.text.strip()
        exps.append([doc,disease,patient,geo_loc,source,exp_time,content,doc_reply])
    return exps


def get_exps_one_url(exp_url):
    set_browser_local(exp_url)
    time.sleep(1)
    exps = get_exps_one_page_bs()
    try:
        total_page_a = get_element(browser,'//div[@class="p_bar"]/a[@class="p_text"][@rel="true"][contains(text(),"共")]')
    #total_page_a = browser.find_element_by_xpath('//div[@class="p_bar"]/a[@class="p_text"][contains(text(),"共")]')
        total_page = total_page_a.text.replace('共','').replace('页','').strip()
        total_page = int(total_page)
    except:
        total_page = 1
    print("total_page = {}".format(total_page))
    for page in range(2,total_page+1):
        exp_url_page = exp_url + '?p={}'.format(page)
        set_browser_local(exp_url_page)
        time.sleep(1)
        exps += get_exps_one_page_bs()
        print("page {}".format(page))
    return exps

def get_experience():
    got_exp_url = load_list(list_got_exp_url)
    with open(list_exp_url) as f:
        for line in f.readlines():
            line = line.strip()
            if line in got_exp_url:
                print('{} already parsed'.format(line))
            else:
                exps = get_exps_one_url(line)
                wb = xl.load_workbook(experience)
                sheet = wb.active
                for ei in exps:
                    sheet.append(ei)
                wb.save(experience)
                write_to_list(list_got_exp_url,line)
                write_log("get {} exps for {}".format(len(exps),line))
#get_experience()


def get_shares_one_page_bs():
    bs_page = bss(browser.page_source,'html.parser')
    try:
        bs_tables = bs_page.find('div',class_='doctorjyjy').find_all('table',class_='doctorjy')
    except:
        return -1
    doc = bs_page.find('div',id='doctor_header').find('h1').text.strip()
    shares = []
    for bs_table in bs_tables:
        bs_tds1 = bs_table.find('table').find_all('td')
        ctype = ''
        patient = ''
        disease = ''
        aim = ''
        treatment = ''
        outcome = ''
        attitude = ''
        write_time = ''
        method = ''
        content = ''
        reason = ''
        channel = ''
        status = ''
        expense = ''
        useful = '0'
        recommendation = '0'
        user_id = ''
        for bs_td in bs_tds1:
            if re.search('患者：',bs_td.text):
                patient = re.sub('患者：','',bs_td.text).strip()
            elif re.search('时间：',bs_td.text):
                write_time = re.sub('时间：','',bs_td.text).strip()
            elif re.search('患者于.*发表',bs_td.text):
                write_time = bs_td.text.strip()
            elif re.search('所患疾病：',bs_td.text):
                disease = re.sub('所患疾病：','',bs_td.text).strip()
            elif re.search('看病目的：',bs_td.text):
                aim = re.sub('看病目的：','',bs_td.text).strip()
            elif re.search('治疗方式：',bs_td.text):
                treatment = re.sub('治疗方式：','',bs_td.text).strip()
            elif re.search('疗效：',bs_td.text):
                outcome = re.sub('疗效：','',bs_td.text).strip()
            elif re.search('态度：',bs_td.text):
                attitude = re.sub('态度：','',bs_td.text).strip()
        bs_td_content = bs_table.find('td',class_='spacejy')
        bs_span_title = bs_td_content.find('span',class_='gray')
        if bs_span_title and re.search('感谢信：',bs_span_title.text):
            ctype='thank'
        else:
            ctype = 'visit'
            bs_method_a = bs_td_content.find('a',class_='orange')
            if bs_method_a:
                method = bs_method_a.text.strip()
        contents = bs_td_content.find_all(text=True,recursive=False)
        contents = [s.strip() for s in contents]
        content = ' '.join(contents).strip()
        bs_tr = bs_td_content.parent.findNext('tr')
        bs_divs = bs_tr.find_all('div',class_='gray')
        for bs_div in bs_divs:
            if re.search('选择该医生就诊的理由：',bs_div.text):
                reason = re.sub('选择该医生就诊的理由：','',bs_div.text).strip()
            elif re.search('本次挂号途径：',bs_div.text):
                channel = re.sub('本次挂号途径：','',bs_div.text).strip()
            elif re.search('目前病情状态：',bs_div.text):
                status = re.sub('目前病情状态：','',bs_div.text).strip()
            elif re.search('本次看病费用总计：',bs_div.text):
                expense = re.sub('本次看病费用总计：','',bs_div.text).strip()
        bs_value_span = ''
        for bs_span in bs_table.find_all('span'):
            if re.search('这条有参考价值吗？',bs_span.text):
                bs_value_span = bs_span.parent.find('span',class_='orange')
                if bs_value_span:
                    recommend_m = re.search('[0-9]+',bs_value_span.text)
                    if recommend_m:
                        recommendation = recommend_m.group()
                        useful = '1'
        for bs_span in bs_table.find_all('span',class_='green'):
            if re.search('回应',bs_span.text):
                user_id = re.sub('回应','',bs_span.text).strip()
                if re.search('此患者',user_id):
                    user_id = ''
                break
        shares.append([doc,ctype,patient,disease,aim,treatment,outcome,attitude,write_time,
                  method,content,reason,channel,status,expense,useful,recommendation,user_id])
    return shares
def get_shares_one_url(share_url):
    set_browser(share_url)
    time.sleep(1)
    update_share_url = browser.current_url
    set_browser_local(update_share_url)
    shares = get_shares_one_page_bs()
    try:
        total_page_a = get_element(browser,'//div[@class="p_bar"]/a[@class="p_text"][@rel="true"][contains(text(),"共")]')
    #total_page_a = browser.find_element_by_xpath('//div[@class="p_bar"]/a[@class="p_text"][contains(text(),"共")]')
        total_page = total_page_a.text.replace('共','').replace('页','').strip()
        total_page = int(total_page)
    except:
        total_page = 1
    print("total {} pages".format(total_page))
    suffixm = re.search('\.[^.]*$',update_share_url)
    if suffixm:
        suffix = suffixm.group()
    else:
        suffix = ''
    base_url = re.sub('\.[^.]*$','',update_share_url)
    for page in range(2,total_page+1):
        share_page_url = base_url + '/{}{}'.format(page,suffix)
        set_browser_local(share_page_url)
        #time.sleep(1)
        page_shares = get_shares_one_page_bs()
        while page_shares == -1:
            set_browser_local(share_page_url,replace=True)
            page_shares = get_shares_one_page_bs()
        shares += page_shares
        print('page {}'.format(page))
    return shares

#x = get_shares_one_url('http://www.haodf.com/doctor/DE4r0BCkuHzdehajeS3cdu-554j7S/jingyan/2.htm')
def get_shares():
    got_share_url = load_list(list_got_share_url)
    with open(list_share_url) as f:
        for line in f.readlines():
            line = line.strip()
            if line in got_share_url:
                print('{} already parsed'.format(line))
            else:
                shares = get_shares_one_url(line)
                wb = xl.load_workbook(sharing)
                sheet = wb.active
                for shi in shares:
                    try:
                        sheet.append(shi)
                    except:
                        for i in range(len(shi)):
                            shi[i] = ILLEGAL_CHARACTERS_RE.sub(r'',shi[i])
                        sheet.append(shi)
                wb.save(sharing)
                write_to_list(list_got_share_url,line)
                write_log("get {} shares for {}".format(len(shares),line))
#get_shares()
browser.close()
